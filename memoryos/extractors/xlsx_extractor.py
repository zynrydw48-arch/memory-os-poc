import openpyxl

from memoryos.extractors.result import ExtractionResult

MAX_TOTAL_CHARS = 20_000  # bounds pathologically large sheets


def extract_xlsx(path) -> ExtractionResult:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    text_parts = []
    total_chars = 0

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                text_parts.append(text)
                total_chars += len(text)
            if total_chars >= MAX_TOTAL_CHARS:
                break
        if total_chars >= MAX_TOTAL_CHARS:
            break

    sheet_names = workbook.sheetnames
    workbook.close()

    return ExtractionResult(
        text="\n".join(text_parts),
        structural_metadata={"sheet_names": sheet_names},
    )
